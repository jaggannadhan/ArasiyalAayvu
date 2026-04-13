"""
ASI (Annual Survey of Industries) — state-level ingestor
Source: MOSPI publications — https://www.mospi.gov.in/publications-reports
        Also: https://www.mospi.gov.in/asi-summary-results

Raw Excel files expected at: data/raw/asi/asi_{year}.xlsx
  e.g. data/raw/asi/asi_2023_24.xlsx

  How to download:
    1. Visit: https://www.mospi.gov.in/asi-summary-results
    2. Find the latest ASI edition (e.g. "ASI 2023-24")
    3. Download "Volume I — Summary Results of Factory Sector"
    4. Open the Excel, locate the state-wise summary sheet
    5. Save the STATE-WISE sheet as:  data/raw/asi/asi_{year}.xlsx
       Year format: "2023_24" for 2023-24, "2022_23" for 2022-23

  OR — for already-published state-wise CSVs on data.gov.in:
    https://www.data.gov.in/catalog/all-india-and-state-wise-annual-survey-industries-factory-sector

  Expected columns (MOSPI state-wise summary table):
    State/UT | No. of Factories | Fixed Capital (₹ Lakh) | Working Capital (₹ Lakh) |
    Total Input (₹ Lakh) | Total Output (₹ Lakh) | Net Value Added (₹ Lakh) |
    Total Emoluments (₹ Lakh) | Workers | Persons Engaged

Outputs: data/processed/asi_ts.json

Run:
    python scrapers/asi_ingest.py               # parse all Excel files, write JSON
    python scrapers/asi_ingest.py --upload      # also upload to Firestore
    python scrapers/asi_ingest.py --probe       # print sheet/column info and exit
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl

BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))
from ts_utils import load_timeseries, upsert_snapshot, save_timeseries, upload_snapshot_to_firestore, get_firestore_client

RAW_DIR  = BASE_DIR / "data" / "raw" / "asi"
OUT_PATH = BASE_DIR / "data" / "processed" / "asi_ts.json"

FOCUS_STATES = {
    "Tamil Nadu", "Kerala", "Karnataka", "Andhra Pradesh", "Telangana",
}

# Column name aliases (lowercase substring match)
COLUMN_ALIASES: dict[str, list[str]] = {
    "factories":        ["no. of factories", "number of factories", "factories"],
    "fixed_capital":    ["fixed capital"],
    "working_capital":  ["working capital"],
    "total_input":      ["total input"],
    "total_output":     ["total output", "gross output"],
    "gva":              ["net value added", "nva", "gross value added", "gva"],
    "emoluments":       ["total emoluments", "emolument", "wages"],
    "workers":          ["no. of workers", "number of workers", "workers"],
    "persons_engaged":  ["persons engaged", "persons employed"],
}

STATE_NAME_ALIASES: dict[str, str] = {
    "andhra pradesh (new)": "Andhra Pradesh",
    "telangana (new)":      "Telangana",
    "a & n islands":        "Ignore",
    "d&n haveli":           "Ignore",
    "daman & diu":          "Ignore",
    "all india":            "All India",
    "total":                "All India",
}


def _canonical_state(raw: str) -> Optional[str]:
    s = raw.strip().strip("*").strip()
    low = s.lower()
    alias = STATE_NAME_ALIASES.get(low)
    if alias == "Ignore":
        return None
    if alias:
        return alias
    return s


def _find_state_sheet(wb: openpyxl.Workbook) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Find the sheet containing state-wise summary data."""
    for name in wb.sheetnames:
        nl = name.lower()
        if any(k in nl for k in ("state", "statewise", "state-wise", "statement")):
            return wb[name]
    # Fall back to first sheet
    return wb.active


def _find_header_row(ws, max_search: int = 20) -> Optional[int]:
    """Find the row index (1-based) containing column headers."""
    for row_idx in range(1, max_search):
        row = [str(ws.cell(row_idx, c).value or "").strip().lower() for c in range(1, 10)]
        row_text = " ".join(row)
        if any(k in row_text for k in ("factories", "capital", "output", "value added")):
            return row_idx
    return None


def _map_columns(ws, header_row: int) -> dict[str, int]:
    """Return {metric_key: col_index_1based}."""
    col_map: dict[str, int] = {}
    for c in range(1, 30):
        cell_val = str(ws.cell(header_row, c).value or "").strip().lower()
        if not cell_val:
            continue
        for key, aliases in COLUMN_ALIASES.items():
            if any(a in cell_val for a in aliases) and key not in col_map:
                col_map[key] = c
                break
    return col_map


def _get_num(ws, row_idx: int, col: int) -> Optional[float]:
    val = ws.cell(row_idx, col).value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("N.A.", "").replace("-", "")
    try:
        return float(s)
    except ValueError:
        return None


def _lakh_to_crore(v: Optional[float]) -> Optional[float]:
    """Convert ₹ lakh to ₹ crore (÷ 100)."""
    return round(v / 100, 2) if v is not None else None


def parse_excel(path: Path, period: str, probe: bool) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, data_only=True)

    if probe:
        print(f"\n  Sheets in {path.name}: {wb.sheetnames}")
        ws = _find_state_sheet(wb)
        if ws:
            hr = _find_header_row(ws)
            print(f"  Header row index: {hr}")
            if hr:
                headers = [str(ws.cell(hr, c).value or "").strip() for c in range(1, 25)]
                print(f"  Headers: {[h for h in headers if h]}")
            # Print first 5 data rows
            for ri in range((hr or 1) + 1, (hr or 1) + 6):
                row = [str(ws.cell(ri, c).value or "").strip() for c in range(1, 8)]
                print(f"  Row {ri}: {row}")
        return {}

    ws = _find_state_sheet(wb)
    if not ws:
        print(f"  ERROR: No state-wise sheet found in {path.name}")
        return {}

    header_row = _find_header_row(ws)
    if not header_row:
        print(f"  ERROR: Could not find header row in {path.name}")
        return {}

    col_map = _map_columns(ws, header_row)
    if not col_map:
        print(f"  ERROR: No recognized columns in {path.name}. Run --probe to inspect.")
        return {}

    # State column: first column
    results: dict[str, dict] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw = str(ws.cell(row_idx, 1).value or "").strip()
        if not raw or raw.isdigit():
            continue

        canonical = _canonical_state(raw)
        if canonical is None:
            continue
        if canonical not in FOCUS_STATES and canonical != "All India":
            continue

        def g(key: str) -> Optional[float]:
            col = col_map.get(key)
            return _get_num(ws, row_idx, col) if col else None

        # MOSPI reports values in ₹ lakh — convert monetary fields to ₹ crore
        results[canonical] = {
            "factories":        g("factories"),
            "workers":          g("workers"),
            "persons_engaged":  g("persons_engaged"),
            "wages_cr":         _lakh_to_crore(g("emoluments")),
            "fixed_capital_cr": _lakh_to_crore(g("fixed_capital")),
            "working_capital_cr": _lakh_to_crore(g("working_capital")),
            "total_output_cr":  _lakh_to_crore(g("total_output")),
            "total_input_cr":   _lakh_to_crore(g("total_input")),
            "gva_cr":           _lakh_to_crore(g("gva")),
        }

    return results


def find_excel_files() -> list[tuple[str, Path]]:
    """Return sorted [(period, path)] for all asi_{year}.xlsx files."""
    pairs = []
    for p in RAW_DIR.glob("asi_*.xlsx"):
        m = re.search(r"asi_(\d{4}_\d{2})\.xlsx", p.name)
        if m:
            raw = m.group(1)  # "2023_24"
            period = raw.replace("_", "-")  # "2023-24"
            # Fix: "2023-24" not "2023-4"
            parts = period.split("-")
            if len(parts[1]) == 2:
                period = f"20{parts[0][:2]}{parts[0][2:]}-{parts[1]}"
            pairs.append((period, p))
    return sorted(pairs)


def print_reminder():
    print()
    print("━" * 65)
    print("  ASI — NO DATA FILES FOUND")
    print("━" * 65)
    print(f"  Expected: {RAW_DIR}/asi_{{year}}.xlsx")
    print("  e.g.:     data/raw/asi/asi_2023_24.xlsx")
    print()
    print("  To download:")
    print("  1. Visit: https://www.mospi.gov.in/publications-reports")
    print("     OR:    https://www.mospi.gov.in/asi-summary-results")
    print("  2. Download 'ASI 2023-24 — Volume I Summary Results'")
    print("  3. Open the Excel and find the State-wise summary sheet")
    print("  4. Save as: data/raw/asi/asi_2023_24.xlsx")
    print()
    print("  OR use data.gov.in state-wise CSV:")
    print("    https://www.data.gov.in/catalog/all-india-and-state-wise-annual-survey-industries-factory-sector")
    print("━" * 65)


def main():
    upload = "--upload" in sys.argv
    probe  = "--probe"  in sys.argv

    files = find_excel_files()
    if not files:
        print_reminder()
        sys.exit(0)

    ts = load_timeseries(OUT_PATH)
    meta = {
        "dataset": "asi",
        "source":  "Annual Survey of Industries (MOSPI)",
        "url":     "https://www.mospi.gov.in/publications-reports",
        "note":    "Factory sector. Monetary values in ₹ crore. Focus states + All India.",
    }

    total_snapshots = 0
    first = True

    for period, path in files:
        print(f"\nParsing {path.name} ({period}) …")
        state_data = parse_excel(path, period, probe)
        if probe:
            continue
        if not state_data:
            continue

        for state, snapshot in sorted(state_data.items()):
            upsert_snapshot(ts, state, period, snapshot, meta=meta if first else None)
            first = False
            total_snapshots += 1
            print(f"  {state:<28} factories={snapshot.get('factories')} "
                  f"gva_cr={snapshot.get('gva_cr')}")

    if not probe:
        save_timeseries(ts, OUT_PATH)
        print(f"\nWrote {OUT_PATH}  ({OUT_PATH.stat().st_size // 1024} KB)")
        print(f"Total snapshots: {total_snapshots}")

        if upload and total_snapshots > 0:
            print("\nUploading to Firestore …")
            db = get_firestore_client()
            count = 0
            for display_name, entity in ts["entities"].items():
                for data_period, snapshot in entity["snapshots"].items():
                    upload_snapshot_to_firestore(db, "asi", display_name, data_period, snapshot)
                    count += 1
            print(f"  Uploaded {count} ASI snapshots to Firestore.")


if __name__ == "__main__":
    main()
