"""
NCRB Crime in India — state-level auto-downloader & ingestor
Source: NCRB Additional Tables → States/UTs
        https://www.ncrb.gov.in/crime-in-india-additional-table?year={year}&category=States/UTs

Fetches state-level Excel files published by NCRB.  No manual download needed.

Available years: 2020, 2021, 2022  (2023 not yet published as of Apr 2026)

Excel file structure (consistent across years and crime types):
  Row 1: Title
  Row 2: Headers (SL | State/UT | {year} | ...)
  Row 3: Sub-headers (Cases Under 156_3 | Cases in Police Station | Cases during year | ...)
  Row 4+: State data
  Col 1: Serial number
  Col 2: State/UT name
  Col 5: Cases Reported during the year (TOTAL) ← this is the metric we extract

Metrics extracted (one Excel file per metric):
  total_ipc_crimes        — State-UT-wise Cases under IPC
  crimes_against_women    — State-UT-wise Cases against Women
  crimes_against_children — State-UT-wise Cases against Children
  crimes_against_sc       — State-UT-wise Cases against SCs
  crimes_against_st       — State-UT-wise Cases against STs

Outputs: data/processed/ncrb_ts.json
Caches raw Excel files in: data/raw/ncrb/

Run:
    python scrapers/ncrb_ingest.py               # download + parse, write JSON
    python scrapers/ncrb_ingest.py --upload      # also upload to Firestore
    python scrapers/ncrb_ingest.py --probe       # list available tables for each year
    python scrapers/ncrb_ingest.py --no-download # use cached Excel files only
"""

from __future__ import annotations

import sys
import warnings
from pathlib import Path
from typing import Optional

import openpyxl
import requests
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore")  # suppress SSL/openpyxl warnings

BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))
from ts_utils import (
    load_timeseries, upsert_snapshot, save_timeseries,
    upload_snapshot_to_firestore, get_firestore_client,
)

RAW_DIR  = BASE_DIR / "data" / "raw" / "ncrb"
OUT_PATH = BASE_DIR / "data" / "processed" / "ncrb_ts.json"

CATEGORY_URL = "https://www.ncrb.gov.in/crime-in-india-additional-table?year={year}&category=States/UTs"
TARGET_YEARS = ["2022", "2021", "2020"]

FOCUS_STATES = {
    "Tamil Nadu", "Kerala", "Karnataka", "Andhra Pradesh", "Telangana",
}

# Keywords in Excel title → metric key.
# Titles must start with "State-UT" to exclude Crimehead-wise / Gender-wise files.
# Listed in order; first match wins.
TABLE_METRICS: list[tuple[str, str, str]] = [
    # (required_prefix,  required_keyword,  metric_key)
    ("state-ut", "under ipc",        "total_ipc_crimes"),
    ("state-ut", "against women",    "crimes_against_women"),
    ("state-ut", "against children", "crimes_against_children"),
    ("state-ut", "against scs",      "crimes_against_sc"),
    ("state-ut", "against sts",      "crimes_against_st"),
]

STATE_ALIASES: dict[str, str] = {
    "total (all india)": "All India",
    "total(all india)":  "All India",
}


def _session() -> requests.Session:
    s = requests.Session()
    s.verify = False
    s.headers["User-Agent"] = "Mozilla/5.0 (compatible; ArasiyalAayvu/1.0)"
    return s


def _canonical_state(raw: str) -> Optional[str]:
    stripped = raw.strip().strip("*").rstrip("0123456789").strip()
    low = stripped.lower()
    alias = STATE_ALIASES.get(low)
    if alias:
        return alias
    if stripped in FOCUS_STATES:
        return stripped
    return None


def _fetch_table_links(year: str, sess: requests.Session) -> list[tuple[str, str]]:
    """
    GET the States/UTs category page for a given year.
    Returns list of (title, xlsx_url) for all rows in the table.
    """
    url = CATEGORY_URL.format(year=year)
    resp = sess.get(url, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    results = []
    for row in soup.select("table tbody tr"):
        cells = row.find_all("td")
        link = row.find("a", href=True)
        if len(cells) >= 2 and link:
            title = cells[1].get_text(strip=True)
            href = link["href"]
            if href.endswith(".xlsx"):
                results.append((title, href))
    return results


def _match_metric(title: str) -> Optional[str]:
    """Return metric key if title matches a known TABLE_METRICS entry."""
    tl = title.lower().strip()
    for prefix, keyword, metric in TABLE_METRICS:
        if tl.startswith(prefix) and keyword in tl:
            return metric
    return None


def _cache_path(year: str, metric: str) -> Path:
    return RAW_DIR / f"ncrb_{year}_{metric}.xlsx"


def _download_excel(url: str, dest: Path, sess: requests.Session) -> bytes:
    if dest.exists():
        return dest.read_bytes()
    resp = sess.get(url, timeout=30)
    resp.raise_for_status()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(resp.content)
    return resp.content


def _parse_excel(data: bytes) -> dict[str, float]:
    """
    Parse a NCRB state-level Excel file.
    Returns {canonical_state: total_cases} for focus states + All India.
    Column 2 = state name, Column 5 = Cases Reported during the year (total).
    """
    wb = openpyxl.load_workbook(__import__("io").BytesIO(data), data_only=True)
    ws = wb.active
    results: dict[str, float] = {}

    for row_idx in range(1, ws.max_row + 1):
        raw_name = str(ws.cell(row_idx, 2).value or "").strip()
        if not raw_name:
            continue
        canonical = _canonical_state(raw_name)
        if not canonical:
            continue

        val = ws.cell(row_idx, 5).value
        if val is None:
            continue
        try:
            results[canonical] = float(str(val).replace(",", ""))
        except (ValueError, TypeError):
            continue

    return results


def process_year(year: str, probe: bool, no_download: bool,
                 sess: requests.Session) -> dict[str, dict[str, float]]:
    """
    For one year: discover Excel links, download, parse.
    Returns {state: {metric: value}}.
    """
    print(f"\n── NCRB {year} ──────────────────────────────────────────")

    # Collect per-metric data: {state: {metric: value}}
    state_data: dict[str, dict[str, float]] = {}

    if no_download:
        # Use cached files if available
        table_links = []
        for _prefix, _keyword, metric in TABLE_METRICS:
            cp = _cache_path(year, metric)
            if cp.exists():
                table_links.append((metric, metric, str(cp)))  # fake title
        if not table_links:
            print(f"  No cached files found in {RAW_DIR}")
            return {}
    else:
        try:
            table_links_raw = _fetch_table_links(year, sess)
        except Exception as e:
            print(f"  ERROR fetching category page: {e}")
            return {}

        if probe:
            print(f"  {len(table_links_raw)} tables found:")
            for title, _href in table_links_raw:
                metric = _match_metric(title)
                print(f"    {'[TARGET]' if metric else '        ':10} {title[:65]}")
            return {}

        table_links = []
        for title, href in table_links_raw:
            metric = _match_metric(title)
            if metric:
                table_links.append((title, metric, href))

    for title, metric, url_or_path in table_links:
        try:
            if no_download:
                data = Path(url_or_path).read_bytes()
            else:
                dest = _cache_path(year, metric)
                data = _download_excel(url_or_path, dest, sess)

            parsed = _parse_excel(data)
            print(f"  {metric:<28} states={len(parsed)}")

            for state, value in parsed.items():
                if state not in state_data:
                    state_data[state] = {}
                state_data[state][metric] = value

        except Exception as e:
            print(f"  ERROR processing {metric}: {e}")

    return state_data


def main() -> None:
    upload      = "--upload"      in sys.argv
    probe       = "--probe"       in sys.argv
    no_download = "--no-download" in sys.argv

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    sess = _session()

    ts = load_timeseries(OUT_PATH)
    meta = {
        "dataset": "ncrb",
        "source":  "NCRB Crime in India — Additional Tables (States/UTs)",
        "url":     "https://www.ncrb.gov.in/crime-in-india-additional-table",
        "note":    "Cases registered during the year. Focus states + All India.",
    }

    total_snapshots = 0
    first = True

    for year in TARGET_YEARS:
        state_data = process_year(year, probe, no_download, sess)
        if probe or not state_data:
            continue

        for state, metrics in sorted(state_data.items()):
            upsert_snapshot(ts, state, year, metrics, meta=meta if first else None)
            first = False
            total_snapshots += 1
            print(f"  {state:<28} ipc={metrics.get('total_ipc_crimes')} "
                  f"women={metrics.get('crimes_against_women')}")

    if not probe:
        save_timeseries(ts, OUT_PATH)
        print(f"\nWrote {OUT_PATH}  ({OUT_PATH.stat().st_size // 1024} KB)")
        print(f"Total snapshots: {total_snapshots}")

        if upload and total_snapshots > 0:
            print("\nUploading to Firestore …")
            db = get_firestore_client()
            count = 0
            for display_name, entity in ts["entities"].items():
                for data_period, snapshot in entity["snapshots"].items():
                    upload_snapshot_to_firestore(db, "ncrb", display_name, data_period, snapshot)
                    count += 1
            print(f"  Uploaded {count} NCRB snapshots to Firestore.")


if __name__ == "__main__":
    main()
